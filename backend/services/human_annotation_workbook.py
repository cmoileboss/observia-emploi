"""Construit et réimporte les classeurs d'annotation humaine."""

from __future__ import annotations

import io
import math
import re
from collections import defaultdict
from collections.abc import Mapping, Sequence
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation

from backend.services.human_annotation_protocol import (
    ANNOTATION_INPUT_COLUMNS,
    ANNOTATOR_COLUMNS,
    VALID_MAIN_CRITERIA,
    VALID_UNCERTAINTY_VALUES,
    read_annotation_csv,
    serialize_annotation_csv,
    validate_completed_annotation,
)


INSTRUCTIONS_SHEET_NAME = "Consignes"
PAIR_ID_COLUMN_INDEX = 3
FIRST_CANDIDATE_ROW = 10
PAIR_MARKER_PREFIX = "PAIR_ID:"
ANNOTATION_MARKER_PREFIX = "ANNOTATION:"
SPLITTABLE_TEXT_FIELDS = frozenset(
    {
        "offre_description",
        "offre_competences",
        "contexte_source_complementaire",
        "certification_activites",
        "certification_competences_attestees",
        "certification_metiers_accessibles",
        "certification_secteurs_activite",
        "certification_prerequis",
        "certification_blocs_competences",
    }
)
MAX_TEXT_CHUNK_CHARACTERS = 1600
MAX_TEXT_CHUNK_VISUAL_LINES = 18
LONG_TEXT_CONTENT_WIDTH = 100

DETAILED_OFFER_FIELDS = (
    ("offre_description", "Description"),
    ("offre_competences", "Compétences détaillées"),
    ("contexte_source_complementaire", "Contexte complémentaire"),
)
CANDIDATE_FIELDS = (
    ("certification_code_rncp", "Code RNCP"),
    ("certification_intitule", "Intitulé"),
    ("certification_niveau", "Niveau"),
    ("certification_activites", "Activités"),
    ("certification_competences_attestees", "Compétences attestées"),
    ("certification_metiers_accessibles", "Métiers accessibles"),
    ("certification_secteurs_activite", "Secteurs"),
    ("certification_prerequis", "Prérequis"),
    ("certification_blocs_competences", "Blocs de compétences"),
    ("score", "Score"),
    ("critere_principal", "Critère principal"),
    ("justification", "Justification"),
    ("incertain", "Incertain"),
)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
LABEL_FILL = PatternFill("solid", fgColor="EAF2F8")
LIGHT_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
TOP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


def _calculate_row_height(
    value: object,
    content_width: int = 120,
    minimum: float = 24,
) -> float:
    """Estime une hauteur lisible selon la longueur et les retours à la ligne."""
    text = "" if value is None else str(value)
    visual_lines = sum(
        max(1, math.ceil(len(line) / content_width))
        for line in text.splitlines() or ("",)
    )
    return min(409, max(minimum, 8 + visual_lines * 15))


def _estimated_visual_lines(value: str, content_width: int) -> int:
    """Estime le nombre de lignes visuelles d'un texte avec retour automatique."""
    return sum(
        max(1, math.ceil(len(line) / content_width))
        for line in value.splitlines() or ("",)
    )


def _split_long_text(value: object) -> tuple[str, ...]:
    """Découpe un texte long entre les mots sans perdre aucun caractère."""
    text = "" if value is None else str(value)
    if not text:
        return ("",)
    segments = re.findall(r"\s+|\S+\s*", text)
    chunks: list[str] = []
    current = ""
    for segment in segments:
        candidate = current + segment
        exceeds_limit = (
            len(candidate) > MAX_TEXT_CHUNK_CHARACTERS
            or _estimated_visual_lines(
                candidate,
                LONG_TEXT_CONTENT_WIDTH,
            ) > MAX_TEXT_CHUNK_VISUAL_LINES
        )
        if current and exceeds_limit:
            chunks.append(current)
            current = segment
        else:
            current = candidate
    if current:
        chunks.append(current)
    if "".join(chunks) != text:
        raise ValueError("Le découpage d'un texte XLSX a altéré son contenu.")
    return tuple(chunks)


def _summarize_offer_competences(value: object, maximum_length: int = 240) -> str:
    """Produit un rappel compact sans altérer les compétences détaillées."""
    text = " ".join(str(value or "").split())
    if len(text) <= maximum_length:
        return text
    return text[: maximum_length - 1].rstrip() + "…"


def _group_rows_by_offer(
    rows: Sequence[Mapping[str, str]],
) -> tuple[tuple[Mapping[str, str], ...], ...]:
    """Regroupe les lignes par offre en conservant leur ordre d'origine."""
    grouped: dict[str, list[Mapping[str, str]]] = defaultdict(list)
    for row in rows:
        offer_group_id = str(row.get("offer_group_id") or "").strip()
        if not offer_group_id:
            raise ValueError("Une ligne CSV ne possède pas de offer_group_id.")
        grouped[offer_group_id].append(row)
    return tuple(tuple(group) for group in grouped.values())


def _add_instructions_sheet(workbook: Workbook) -> None:
    """Ajoute la feuille de consignes destinée aux annotateurs."""
    worksheet = workbook.active
    worksheet.title = INSTRUCTIONS_SHEET_NAME
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 24
    for column in ("B", "C", "D", "E", "F", "G"):
        worksheet.column_dimensions[column].width = 16
    worksheet.merge_cells("A1:G1")
    title = worksheet["A1"]
    title.value = "Annotation offre–certification RNCP"
    title.fill = HEADER_FILL
    title.font = Font(color="FFFFFF", bold=True, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28
    instructions = (
        "Chaque feuille Offre regroupe les certifications candidates d'une seule offre.",
        "Remplir uniquement les cellules jaunes : score, critère principal, justification et incertain.",
        "Score : 0 à 3. Incertain : OUI ou NON.",
        "Ne pas modifier les intitulés, descriptions ou informations de certification.",
        "Le pair_id est conservé dans une colonne masquée pour sécuriser la réimportation.",
    )
    for row_index, instruction in enumerate(instructions, start=3):
        worksheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=7,
        )
        cell = worksheet.cell(row=row_index, column=1, value=instruction)
        cell.alignment = TOP_ALIGNMENT
        worksheet.row_dimensions[row_index].height = 30
    worksheet.protection.sheet = True


def _add_list_validation(
    worksheet,
    cell,
    values: Sequence[str],
) -> None:
    """Ajoute une liste déroulante explicite à une cellule éditable."""
    validation = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
    )
    validation.error = "Choisir une valeur dans la liste."
    validation.errorTitle = "Valeur invalide"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(cell)


def _style_offer_sheet(worksheet) -> None:
    """Applique la mise en page commune d'une feuille d'offre."""
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85
    worksheet.freeze_panes = "A5"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 120
    worksheet.column_dimensions["C"].hidden = True
    worksheet.column_dimensions["C"].width = 28


def _write_offer_header(
    worksheet,
    first_row: Mapping[str, str],
) -> int:
    """Écrit le contexte variable et retourne la ligne suivante disponible."""
    worksheet.merge_cells("A1:B1")
    title = worksheet["A1"]
    title.value = f"Offre — {first_row.get('offre_intitule', '')}"
    title.fill = HEADER_FILL
    title.font = Font(color="FFFFFF", bold=True, size=13)
    title.alignment = TOP_ALIGNMENT
    worksheet.row_dimensions[1].height = 26

    compact_rows = (
        (2, "Appellation", first_row.get("offre_appellation", "")),
        (
            3,
            "ROME",
            " — ".join(
                value
                for value in (
                    first_row.get("offre_code_rome", ""),
                    first_row.get("offre_libelle_rome", ""),
                )
                if value
            ),
        ),
        (
            4,
            "Résumé compétences",
            _summarize_offer_competences(first_row.get("offre_competences", "")),
        ),
    )
    for row_index, label, value in compact_rows:
        label_cell = worksheet.cell(row=row_index, column=1, value=label)
        label_cell.fill = LABEL_FILL
        label_cell.font = Font(bold=True)
        label_cell.border = LIGHT_BORDER
        label_cell.alignment = TOP_ALIGNMENT
        value_cell = worksheet.cell(
            row=row_index,
            column=2,
            value=value,
        )
        value_cell.border = LIGHT_BORDER
        value_cell.alignment = TOP_ALIGNMENT
        worksheet.row_dimensions[row_index].height = _calculate_row_height(
            value,
            minimum=24,
        )

    worksheet.merge_cells("A5:B5")
    section = worksheet["A5"]
    section.value = "Contexte détaillé de l'offre"
    section.fill = SECTION_FILL
    section.font = Font(bold=True)
    section.alignment = TOP_ALIGNMENT
    row_index = 6
    for field, label in DETAILED_OFFER_FIELDS:
        value = first_row.get(field, "")
        chunks = (
            _split_long_text(value)
            if field in SPLITTABLE_TEXT_FIELDS
            else (value,)
        )
        for chunk_index, chunk in enumerate(chunks, start=1):
            visible_label = (
                f"{label} ({chunk_index}/{len(chunks)})"
                if len(chunks) > 1
                else label
            )
            label_cell = worksheet.cell(
                row=row_index,
                column=1,
                value=visible_label,
            )
            label_cell.fill = LABEL_FILL
            label_cell.font = Font(bold=True)
            label_cell.border = LIGHT_BORDER
            label_cell.alignment = TOP_ALIGNMENT
            value_cell = worksheet.cell(row=row_index, column=2, value=chunk)
            value_cell.border = LIGHT_BORDER
            value_cell.alignment = TOP_ALIGNMENT
            worksheet.row_dimensions[row_index].height = _calculate_row_height(
                chunk,
                content_width=LONG_TEXT_CONTENT_WIDTH,
                minimum=42,
            )
            worksheet.row_dimensions[row_index].outlineLevel = 1
            worksheet.row_dimensions[row_index].hidden = False
            row_index += 1
    return row_index + 1


def _write_candidate_block(
    worksheet,
    candidate_row: Mapping[str, str],
    block_index: int,
    start_row: int,
) -> int:
    """Écrit un bloc de hauteur variable et retourne la ligne suivante."""
    worksheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=2,
    )
    code_rncp = candidate_row.get("certification_code_rncp", "")
    certification_title = candidate_row.get("certification_intitule", "")
    header = worksheet.cell(
        row=start_row,
        column=1,
        value=(
            f"Certification {block_index + 1:02d} — "
            f"{code_rncp} — {certification_title}"
        ),
    )
    header.fill = SECTION_FILL
    header.font = Font(bold=True, color="1F1F1F")
    header.alignment = TOP_ALIGNMENT
    worksheet.row_dimensions[start_row].height = _calculate_row_height(
        header.value,
        content_width=140,
        minimum=28,
    )
    worksheet.cell(
        row=start_row,
        column=PAIR_ID_COLUMN_INDEX,
        value=PAIR_MARKER_PREFIX + str(candidate_row.get("pair_id", "")),
    )
    pair_id = str(candidate_row.get("pair_id", ""))
    row_index = start_row + 1
    for field, label in CANDIDATE_FIELDS:
        value = candidate_row.get(field, "")
        chunks = (
            _split_long_text(value)
            if field in SPLITTABLE_TEXT_FIELDS
            else (value,)
        )
        for chunk_index, chunk in enumerate(chunks, start=1):
            visible_label = (
                f"{label} ({chunk_index}/{len(chunks)})"
                if len(chunks) > 1
                else label
            )
            label_cell = worksheet.cell(
                row=row_index,
                column=1,
                value=visible_label,
            )
            label_cell.fill = LABEL_FILL
            label_cell.font = Font(bold=True)
            label_cell.border = LIGHT_BORDER
            label_cell.alignment = TOP_ALIGNMENT
            value_cell = worksheet.cell(row=row_index, column=2, value=chunk)
            value_cell.border = LIGHT_BORDER
            value_cell.alignment = TOP_ALIGNMENT
            if field in ANNOTATION_INPUT_COLUMNS:
                value_cell.fill = EDITABLE_FILL
                worksheet.cell(
                    row=row_index,
                    column=PAIR_ID_COLUMN_INDEX,
                    value=f"{ANNOTATION_MARKER_PREFIX}{field}:{pair_id}",
                )
            minimum_height = (
                42
                if field in SPLITTABLE_TEXT_FIELDS | {"justification"}
                else 24
            )
            content_width = (
                LONG_TEXT_CONTENT_WIDTH
                if field in SPLITTABLE_TEXT_FIELDS
                else 120
            )
            worksheet.row_dimensions[row_index].height = _calculate_row_height(
                chunk,
                content_width=content_width,
                minimum=minimum_height,
            )
            if field == "score":
                _add_list_validation(
                    worksheet,
                    value_cell,
                    ("0", "1", "2", "3"),
                )
            elif field == "critere_principal":
                _add_list_validation(
                    worksheet,
                    value_cell,
                    VALID_MAIN_CRITERIA,
                )
            elif field == "incertain":
                _add_list_validation(
                    worksheet,
                    value_cell,
                    VALID_UNCERTAINTY_VALUES,
                )
            row_index += 1
    return row_index + 1


def build_annotation_workbook(template_csv: bytes) -> bytes:
    """Construit le classeur lisible compagnon d'un CSV annotateur."""
    rows = read_annotation_csv(template_csv, "CSV modèle du classeur")
    if not rows:
        raise ValueError("Le CSV modèle du classeur est vide.")
    if tuple(rows[0]) != ANNOTATOR_COLUMNS:
        raise ValueError("Les colonnes du CSV modèle sont invalides.")
    pair_ids = [str(row.get("pair_id") or "").strip() for row in rows]
    if not all(pair_ids) or len(set(pair_ids)) != len(pair_ids):
        raise ValueError("Les pair_id du CSV modèle sont invalides ou dupliqués.")

    workbook = Workbook()
    _add_instructions_sheet(workbook)
    for offer_index, offer_rows in enumerate(_group_rows_by_offer(rows), start=1):
        worksheet = workbook.create_sheet(f"Offre {offer_index:02d}")
        _style_offer_sheet(worksheet)
        next_row = _write_offer_header(worksheet, offer_rows[0])
        for block_index, candidate_row in enumerate(offer_rows):
            next_row = _write_candidate_block(
                worksheet,
                candidate_row,
                block_index,
                next_row,
            )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _read_workbook_responses(workbook_content: bytes) -> dict[str, dict[str, str]]:
    """Lit les quatre réponses de chaque pair_id masqué dans le classeur."""
    try:
        workbook = load_workbook(io.BytesIO(workbook_content), data_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValueError("Le classeur XLSX est invalide ou illisible.") from exc
    declared_pair_ids: set[str] = set()
    responses: dict[str, dict[str, str]] = defaultdict(dict)
    for worksheet in workbook.worksheets:
        if worksheet.title == INSTRUCTIONS_SHEET_NAME:
            continue
        for row_index in range(1, worksheet.max_row + 1):
            marker_value = worksheet.cell(
                row=row_index,
                column=PAIR_ID_COLUMN_INDEX,
            ).value
            if marker_value is None:
                continue
            marker = str(marker_value).strip()
            if marker.startswith(PAIR_MARKER_PREFIX):
                pair_id = marker.removeprefix(PAIR_MARKER_PREFIX).strip()
                if not pair_id or pair_id in declared_pair_ids:
                    raise ValueError(
                        "Le classeur contient un pair_id invalide ou dupliqué."
                    )
                declared_pair_ids.add(pair_id)
                continue
            if marker.startswith(ANNOTATION_MARKER_PREFIX):
                payload = marker.removeprefix(ANNOTATION_MARKER_PREFIX)
                field, separator, pair_id = payload.partition(":")
                if (
                    not separator
                    or field not in ANNOTATION_INPUT_COLUMNS
                    or not pair_id
                    or field in responses[pair_id]
                ):
                    raise ValueError(
                        "Le classeur contient un marqueur d'annotation invalide "
                        "ou dupliqué."
                    )
                responses[pair_id][field] = _annotation_value(
                    worksheet.cell(row=row_index, column=2).value
                )
                continue
            raise ValueError("Le classeur contient un marqueur technique inconnu.")

    if set(responses) != declared_pair_ids:
        raise ValueError("Les réponses ne correspondent pas aux pair_id du classeur.")
    expected_fields = set(ANNOTATION_INPUT_COLUMNS)
    for pair_id, response in responses.items():
        if set(response) != expected_fields:
            raise ValueError(
                f"Les quatre réponses sont incomplètes pour {pair_id}."
            )
    return dict(responses)


def _annotation_value(value: object) -> str:
    """Convertit une réponse Excel en texte compatible avec le CSV canonique."""
    return "" if value is None else str(value)


def import_workbook_annotations(
    template_csv: bytes,
    workbook_content: bytes,
) -> bytes:
    """Réinjecte les réponses XLSX dans le CSV modèle puis les valide."""
    template_rows = read_annotation_csv(template_csv, "CSV modèle original")
    if not template_rows or tuple(template_rows[0]) != ANNOTATOR_COLUMNS:
        raise ValueError("Le CSV modèle original est vide ou invalide.")
    template_by_pair = {
        str(row.get("pair_id") or "").strip(): row
        for row in template_rows
    }
    if "" in template_by_pair or len(template_by_pair) != len(template_rows):
        raise ValueError("Le CSV modèle contient un pair_id invalide ou dupliqué.")
    responses = _read_workbook_responses(workbook_content)
    if set(responses) != set(template_by_pair):
        raise ValueError(
            "Les pair_id du classeur ne correspondent pas exactement au CSV modèle."
        )

    completed_rows = []
    for template_row in template_rows:
        completed_row = dict(template_row)
        completed_row.update(responses[template_row["pair_id"]])
        completed_rows.append(completed_row)
    completed_csv = serialize_annotation_csv(completed_rows)
    validate_completed_annotation(template_csv, completed_csv)
    return completed_csv
