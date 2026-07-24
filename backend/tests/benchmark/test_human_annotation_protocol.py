"""Tests métier du protocole d'annotation humaine du lot 6A."""

import csv
import io
import json
from collections import defaultdict

import pytest
from openpyxl import load_workbook

from backend.services.human_annotation_protocol import (
    ANNOTATOR_COLUMNS,
    HIDDEN_FIELDS,
    PROTECTED_COLUMNS,
    VALID_MAIN_CRITERIA,
    AnnotationPackageResult,
    Lot5AnnotationInputs,
    audit_cross_split_duplicates,
    build_annotation_packages,
    build_offer_group_id,
    build_pair_id,
    build_source_complementary_context,
    export_annotation_packages,
    load_lot5_annotation_inputs,
    normalize_duplicate_text,
    select_pilot_offer_keys,
    validate_completed_annotation,
)
from backend.services.human_annotation_workbook import (
    ANNOTATION_MARKER_PREFIX,
    INSTRUCTIONS_SHEET_NAME,
    PAIR_ID_COLUMN_INDEX,
    PAIR_MARKER_PREFIX,
    build_annotation_workbook,
    import_workbook_annotations,
)


def make_offer(index: int, split: str) -> dict:
    has_skills = index % 2 == 0
    has_requirement = index % 3 == 0
    richness = ("COURTE", "MOYENNE", "LONGUE")[index % 3]
    description_length = {"COURTE": 300, "MOYENNE": 700, "LONGUE": 1600}[
        richness
    ]
    return {
        "source": "FRANCE_TRAVAIL",
        "source_offer_id": f"FT-{index:03d}",
        "database_offer_id": index,
        "split": split,
        "code_rome": f"M18{index % 5:02d}",
        "richesse": {
            "description": richness,
            "competences_structurees": has_skills,
            "exigence_france_travail": has_requirement,
        },
        "champs_sources": {
            "intitule": f"Offre unique {index}",
            "appellation": f"Appellation {index % 4}",
            "libelle_rome": f"Métier ROME {index % 5}",
            "description": f"Description unique {index} "
            + (f"contenu-{index} " * description_length),
            "competences": (
                [{"code": "C1", "libelle": f"Compétence {index}"}]
                if has_skills
                else []
            ),
            "exigences_france_travail": (
                [
                    {
                        "intitule": "Diplôme informatique",
                        "code_source": "BAC+3",
                        "niveau": "6",
                        "commentaire": "Souhaité",
                    }
                ]
                if has_requirement
                else []
            ),
        },
    }


def make_candidate(index: int) -> dict:
    return {
        "code_rncp": f"RNCP{index:03d}",
        "intitule_officiel": f"Certification {index}",
        "raison_selection": (
            "ROME_DIRECT", "ROME_PROCHE", "NEGATIF_CONTROLE"
        )[index % 3],
        "donnees_officielles": {
            "niveau": {"code": "6", "libelle": "Niveau 6"},
            "activites_visees": f"Activités {index}",
            "competences_attestees": f"Compétences attestées {index}",
            "metiers_accessibles": f"Métier {index}",
            "secteurs_activite": "Numérique",
            "prerequis": "Baccalauréat",
            "blocs_competences": [
                {
                    "code": f"BC{index:02d}",
                    "libelle": f"Bloc {index}",
                    "competences": f"Compétences du bloc {index}",
                }
            ],
        },
    }


def make_inputs() -> Lot5AnnotationInputs:
    offers = tuple(
        make_offer(index, "development" if index <= 40 else "validation")
        for index in range(1, 61)
    )
    candidates = tuple(make_candidate(index) for index in range(1, 13))
    pools = tuple(
        {
            "source": offer["source"],
            "source_offer_id": offer["source_offer_id"],
            "database_offer_id": offer["database_offer_id"],
            "split": offer["split"],
            "code_rome": offer["code_rome"],
            "candidats": candidates,
        }
        for offer in offers
    )
    pairs = frozenset(
        (offer["source"], offer["source_offer_id"], candidate["code_rncp"])
        for offer in offers
        for candidate in candidates
    )
    return Lot5AnnotationInputs(
        manifest={
            "format_version": "observia-offre-certification-sample-v1",
            "compteurs": {"offres": 60, "couples_offre_certification": 720},
        },
        offers=offers,
        candidate_pools=pools,
        annotation_pairs=pairs,
        artifact_sha256={name: f"hash-{name}" for name in (
            "sample_manifest.json",
            "evaluation_offers.json",
            "candidate_pools.jsonl",
            "annotation_template.csv",
        )},
    )


def make_small_multisource_inputs() -> Lot5AnnotationInputs:
    offers = [
        make_offer(index, "development" if index <= 5 else "validation")
        for index in range(1, 8)
    ]
    offers[0] = {
        **offers[0],
        "source": "FREE_WORK",
        "source_offer_id": "FW-001",
        "database_offer_id": None,
        "code_rome": "",
        "champs_sources": {
            **offers[0]["champs_sources"],
            "appellation": "",
            "libelle_rome": "",
            "exigences_france_travail": [],
        },
    }
    offers[1] = {
        **offers[1],
        "source": "FREE_WORK",
        "source_offer_id": "FW-002",
        "database_offer_id": None,
        "champs_sources": {
            **offers[1]["champs_sources"],
            "exigences_france_travail": [],
            "contexte_source_complementaire": "Contexte Free-Work fourni.",
        },
    }
    candidates = tuple(make_candidate(index) for index in range(1, 4))
    pools = tuple(
        {
            "source": offer["source"],
            "source_offer_id": offer["source_offer_id"],
            "database_offer_id": offer["database_offer_id"],
            "split": offer["split"],
            "code_rome": offer["code_rome"],
            "candidats": candidates,
        }
        for offer in offers
    )
    pairs = frozenset(
        (offer["source"], offer["source_offer_id"], candidate["code_rncp"])
        for offer in offers
        for candidate in candidates
    )
    return Lot5AnnotationInputs(
        manifest={
            "format_version": "observia-offre-certification-sample-v1",
            "parametres": {"candidate_pool_size": 3},
            "compteurs": {"offres": 7, "couples_offre_certification": 21},
        },
        offers=tuple(offers),
        candidate_pools=pools,
        annotation_pairs=pairs,
        artifact_sha256={name: f"hash-{name}" for name in (
            "sample_manifest.json",
            "evaluation_offers.json",
            "candidate_pools.jsonl",
            "annotation_template.csv",
        )},
    )


@pytest.fixture(scope="module")
def package_result() -> AnnotationPackageResult:
    return build_annotation_packages(make_inputs())


def read_csv(content: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))


def write_csv(rows: list[dict[str, str]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=ANNOTATOR_COLUMNS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def complete_rows(content: bytes) -> list[dict[str, str]]:
    rows = read_csv(content)
    for row in rows:
        row["score"] = "2"
        row["critere_principal"] = VALID_MAIN_CRITERIA[0]
        row["justification"] = "Justification humaine suffisamment explicite."
        row["incertain"] = "NON"
    return rows


def orders_by_group(content: bytes) -> dict[str, tuple[str, ...]]:
    grouped: dict[str, list[str]] = defaultdict(list)
    for row in read_csv(content):
        grouped[row["offer_group_id"]].append(row["certification_code_rncp"])
    return {key: tuple(values) for key, values in grouped.items()}


def test_normalization_and_exact_cross_split_duplicate_audit():
    development = make_offer(1, "development")
    validation = make_offer(2, "validation")
    development["champs_sources"]["intitule"] = "Développeur  Python"
    validation["champs_sources"]["intitule"] = "developpeur-python"
    development["champs_sources"]["description"] = "Même description"
    validation["champs_sources"]["description"] = "meme   description"

    audit = audit_cross_split_duplicates((development, validation))

    assert normalize_duplicate_text(" ÉTUDES---SI ") == "etudes si"
    assert len(audit.exact_duplicates) == 1
    assert not audit.near_duplicates


def test_conservative_near_duplicate_audit():
    development = make_offer(1, "development")
    validation = make_offer(2, "validation")
    development["champs_sources"]["intitule"] = "Ingénieur système"
    validation["champs_sources"]["intitule"] = "ingenieur systeme"
    common_tokens = [f"mot{index}" for index in range(100)]
    development["champs_sources"]["description"] = " ".join(common_tokens)
    validation["champs_sources"]["description"] = " ".join(
        (*common_tokens, "complement")
    )

    audit = audit_cross_split_duplicates((development, validation))

    assert not audit.exact_duplicates
    assert len(audit.near_duplicates) == 1
    assert audit.near_duplicates[0].token_jaccard >= 0.95


def test_generation_stops_on_cross_split_duplicate():
    inputs = make_inputs()
    offers = [dict(offer) for offer in inputs.offers]
    offers[40] = {
        **offers[40],
        "champs_sources": dict(offers[0]["champs_sources"]),
    }
    duplicated_inputs = Lot5AnnotationInputs(
        manifest=inputs.manifest,
        offers=tuple(offers),
        candidate_pools=inputs.candidate_pools,
        annotation_pairs=inputs.annotation_pairs,
        artifact_sha256=inputs.artifact_sha256,
    )

    with pytest.raises(ValueError, match="Doublons inter-splits"):
        build_annotation_packages(duplicated_inputs)


def test_pair_and_group_identifiers_are_stable():
    assert build_pair_id("FRANCE_TRAVAIL", "FT-1", "RNCP1") == build_pair_id(
        "FRANCE_TRAVAIL", "FT-1", "RNCP1"
    )
    assert build_pair_id("FRANCE_TRAVAIL", "FT-1", "RNCP1") != build_pair_id(
        "FREE_WORK", "FT-1", "RNCP1"
    )
    assert build_offer_group_id("FRANCE_TRAVAIL", "FT-1").startswith("offer_")


def test_annotators_have_same_candidates_in_different_reproducible_orders(
    package_result,
):
    first = orders_by_group(package_result.artifacts["pilot_annotator_1.csv"])
    second = orders_by_group(package_result.artifacts["pilot_annotator_2.csv"])

    assert set(first) == set(second)
    assert all(set(first[key]) == set(second[key]) for key in first)
    assert all(first[key] != second[key] for key in first)
    repeated = build_annotation_packages(make_inputs())
    assert repeated.artifacts["pilot_annotator_1.csv"] == (
        package_result.artifacts["pilot_annotator_1.csv"]
    )


def test_packets_hide_reference_fields_and_separate_splits(package_result):
    pilot_rows = read_csv(package_result.artifacts["pilot_annotator_1.csv"])
    development_rows = read_csv(
        package_result.artifacts["development_annotator_1.csv"]
    )
    validation_rows = read_csv(
        package_result.artifacts["validation_annotator_1.csv"]
    )

    assert len(pilot_rows) == 60
    assert len(development_rows) == 480
    assert len(validation_rows) == 240
    assert not (set(HIDDEN_FIELDS) & set(pilot_rows[0]))
    assert {
        row["offer_group_id"] for row in development_rows
    }.isdisjoint({row["offer_group_id"] for row in validation_rows})


def test_formats_france_travail_context_in_the_generic_column(package_result):
    offer = make_offer(3, "development")
    expected_context = "Diplôme informatique — BAC+3 — 6 — Souhaité"
    pilot_rows = read_csv(package_result.artifacts["pilot_annotator_1.csv"])

    assert build_source_complementary_context(
        offer["source"], offer["champs_sources"]
    ) == expected_context
    assert "contexte_exigences_france_travail" not in ANNOTATOR_COLUMNS
    assert "source" not in ANNOTATOR_COLUMNS
    assert any(
        row["contexte_source_complementaire"] == expected_context
        for row in pilot_rows
    )


def test_csv_bom_preserves_french_text_and_accepts_legacy_utf8():
    inputs = make_inputs()
    offer = {
        **inputs.offers[0],
        "champs_sources": {
            **inputs.offers[0]["champs_sources"],
            "intitule": "développeur réseau",
            "description": "sécurité et réseau",
            "competences": [{"code": "C1", "libelle": "compétences"}],
        },
    }
    enriched_inputs = Lot5AnnotationInputs(
        manifest=inputs.manifest,
        offers=(offer, *inputs.offers[1:]),
        candidate_pools=inputs.candidate_pools,
        annotation_pairs=inputs.annotation_pairs,
        artifact_sha256=inputs.artifact_sha256,
    )
    result = build_annotation_packages(enriched_inputs)
    template = result.artifacts["development_annotator_1.csv"]

    assert template.startswith(b"\xef\xbb\xbf")
    rows = read_csv(template)
    assert tuple(rows[0]) == ANNOTATOR_COLUMNS
    offer_rows = [row for row in rows if row["offre_intitule"] == "développeur réseau"]
    assert offer_rows
    assert offer_rows[0]["offre_description"] == "sécurité et réseau"
    assert offer_rows[0]["offre_competences"] == "C1 — compétences"

    completed = write_csv(complete_rows(template))
    validate_completed_annotation(template, completed)

    legacy_template = template[3:]
    legacy_completed = completed[3:]
    validate_completed_annotation(legacy_template, legacy_completed)


def test_formats_multiple_france_travail_requirements_and_ignores_generic_values():
    source_fields = {
        "contexte_source_complementaire": "  ",
        "exigences_france_travail": [
            {
                "intitule": "Diplôme informatique",
                "code_source": "BAC+3",
                "niveau": "6",
                "commentaire": "Souhaité",
            },
            {
                "intitule": "Certification cloud",
                "code_source": "",
                "niveau": None,
                "commentaire": "",
            },
        ],
    }

    assert build_source_complementary_context(
        "FRANCE_TRAVAIL", source_fields
    ) == (
        "Diplôme informatique — BAC+3 — 6 — Souhaité"
        " | Certification cloud"
    )
    for invalid_context in (["invalide"], {"texte": "invalide"}):
        assert build_source_complementary_context(
            "FREE_WORK", {"contexte_source_complementaire": invalid_context}
        ) == ""


def test_pilot_selection_is_deterministic_and_covers_requested_dimensions():
    inputs = make_inputs()
    first = select_pilot_offer_keys(inputs.offers)
    second = select_pilot_offer_keys(reversed(inputs.offers))
    offers_by_key = {
        (offer["source"], offer["source_offer_id"]): offer
        for offer in inputs.offers
    }
    pilot = [offers_by_key[key] for key in first]

    assert first == second
    assert len(first) == 5
    assert all(offer["split"] == "development" for offer in pilot)
    assert len({offer["code_rome"] for offer in pilot}) > 1
    assert len({offer["richesse"]["description"] for offer in pilot}) == 3
    assert {bool(offer["champs_sources"]["competences"]) for offer in pilot} == {
        False,
        True,
    }
    assert {
        bool(
            build_source_complementary_context(
                offer["source"], offer["champs_sources"]
            )
        )
        for offer in pilot
    } == {False, True}


def test_all_generated_artifacts_are_byte_deterministic(package_result):
    second = build_annotation_packages(make_inputs())
    assert package_result.artifacts == second.artifacts
    assert set(package_result.artifacts) == {
        "annotation_manifest.json",
        "annotation_reference.jsonl",
        "pilot_annotator_1.csv",
        "pilot_annotator_2.csv",
        "development_annotator_1.csv",
        "development_annotator_2.csv",
        "validation_annotator_1.csv",
        "validation_annotator_2.csv",
    }


def test_builds_readable_pilot_workbook_with_hidden_pair_ids(package_result):
    workbook_content = build_annotation_workbook(
        package_result.artifacts["pilot_annotator_1.csv"]
    )
    workbook = load_workbook(io.BytesIO(workbook_content))

    assert workbook.sheetnames == [
        INSTRUCTIONS_SHEET_NAME,
        "Offre 01",
        "Offre 02",
        "Offre 03",
        "Offre 04",
        "Offre 05",
    ]
    pair_ids = []
    first_csv_row = read_csv(
        package_result.artifacts["pilot_annotator_1.csv"]
    )[0]
    for worksheet in workbook.worksheets[1:]:
        worksheet_pair_ids = [
            cell.value.removeprefix(PAIR_MARKER_PREFIX)
            for cell in worksheet["C"]
            if isinstance(cell.value, str)
            and cell.value.startswith(PAIR_MARKER_PREFIX)
        ]
        pair_ids.extend(worksheet_pair_ids)
        assert len(worksheet_pair_ids) == 12
        assert worksheet.freeze_panes == "A5"
        assert not worksheet.protection.sheet
        assert worksheet.sheet_view.zoomScale == 85
        assert 110 <= worksheet.column_dimensions["B"].width <= 130
        assert worksheet.column_dimensions["C"].hidden
        validations = worksheet.data_validations.dataValidation
        assert len(validations) == 36
        assert {validation.formula1 for validation in validations} == {
            '"0,1,2,3"',
            '"' + ",".join(VALID_MAIN_CRITERIA) + '"',
            '"OUI,NON"',
        }
        visible_values = {
            str(cell.value)
            for row in worksheet.iter_rows(min_col=1, max_col=2)
            for cell in row
            if cell.value is not None
        }
        assert not (set(HIDDEN_FIELDS) & visible_values)
        assert "annotation_reference.jsonl" not in visible_values
        assert any("Compétences" in value for value in visible_values)
        description_row = next(
            cell.row
            for cell in worksheet["A"]
            if isinstance(cell.value, str)
            and cell.value.startswith("Description")
        )
        appellation_row = next(
            cell.row for cell in worksheet["A"] if cell.value == "Appellation"
        )
        assert description_row > 4
        assert worksheet.row_dimensions[description_row].outlineLevel == 1
        assert not worksheet.row_dimensions[description_row].hidden
        assert worksheet.row_dimensions[description_row].height > (
            worksheet.row_dimensions[appellation_row].height
        )
        long_text_labels = {
            "Description",
            "Compétences détaillées",
            "Contexte complémentaire",
            "Activités",
            "Compétences attestées",
            "Blocs de compétences",
            "Justification",
        }
        long_text_rows = {
            cell.row
            for cell in worksheet["A"]
            if isinstance(cell.value, str)
            and any(cell.value.startswith(label) for label in long_text_labels)
        }
        assert all(
            not (
                merged.min_col <= 2 <= merged.max_col
                and merged.min_row in long_text_rows
            )
            for merged in worksheet.merged_cells.ranges
        )
        first_header_row = next(
            cell.row
            for cell in worksheet["C"]
            if isinstance(cell.value, str)
            and cell.value.startswith(PAIR_MARKER_PREFIX)
        )
        first_header = worksheet.cell(row=first_header_row, column=1).value
        assert first_header.startswith("Certification 01 — RNCP")
    first_sheet = workbook["Offre 01"]
    first_description_row = next(
        cell.row
        for cell in first_sheet["A"]
        if isinstance(cell.value, str)
        and cell.value.startswith("Description")
    )
    first_description_rows = [
        cell.row
        for cell in first_sheet["A"]
        if isinstance(cell.value, str)
        and cell.value.startswith("Description")
    ]
    assert "".join(
        first_sheet.cell(row=row_index, column=2).value or ""
        for row_index in first_description_rows
    ) == first_csv_row["offre_description"]
    assert len(pair_ids) == 60
    assert len(set(pair_ids)) == 60


def test_workbooks_keep_each_annotator_candidate_order(package_result):
    workbook_pair_orders = []
    csv_pair_orders = []
    for annotator in ("annotator_1", "annotator_2"):
        csv_content = package_result.artifacts[f"pilot_{annotator}.csv"]
        csv_pair_orders.append([row["pair_id"] for row in read_csv(csv_content)])
        workbook = load_workbook(
            io.BytesIO(build_annotation_workbook(csv_content)),
            read_only=True,
        )
        workbook_pair_orders.append(
            [
                row[0].value
                for worksheet in workbook.worksheets[1:]
                for row in worksheet.iter_rows(
                    min_col=PAIR_ID_COLUMN_INDEX,
                    max_col=PAIR_ID_COLUMN_INDEX,
                )
                if isinstance(row[0].value, str)
                and row[0].value.startswith(PAIR_MARKER_PREFIX)
            ]
        )

        workbook_pair_orders[-1] = [
            marker.removeprefix(PAIR_MARKER_PREFIX)
            for marker in workbook_pair_orders[-1]
        ]

    assert workbook_pair_orders == csv_pair_orders
    assert set(workbook_pair_orders[0]) == set(workbook_pair_orders[1])
    assert workbook_pair_orders[0] != workbook_pair_orders[1]


def test_splits_very_long_rncp_texts_without_altering_content(package_result):
    rows = read_csv(package_result.artifacts["pilot_annotator_1.csv"])
    activities = (
        "Analyser les besoins réseau et sécurité. " * 140
        + "\n\nPiloter les activités avec les équipes. " * 30
    )
    competences = (
        "Développer des compétences techniques et documenter les résultats. "
        * 90
    )
    blocks = (
        "Bloc de compétences : concevoir, sécuriser et maintenir le système. "
        * 320
    )
    assert len(activities) >= 5_000
    assert len(competences) >= 5_000
    assert len(blocks) >= 20_000
    rows[0]["certification_activites"] = activities
    rows[0]["certification_competences_attestees"] = competences
    rows[0]["certification_blocs_competences"] = blocks

    workbook = load_workbook(
        io.BytesIO(build_annotation_workbook(write_csv(rows)))
    )
    worksheet = workbook["Offre 01"]
    expected_by_label = {
        "Activités": activities,
        "Compétences attestées": competences,
        "Blocs de compétences": blocks,
    }
    for label, expected_text in expected_by_label.items():
        continuation_rows = [
            cell.row
            for cell in worksheet["A"]
            if isinstance(cell.value, str)
            and cell.value.startswith(f"{label} (")
        ]
        chunks = [
            worksheet.cell(row=row_index, column=2).value
            for row_index in continuation_rows
        ]
        assert len(chunks) > 1
        assert "".join(chunks) == expected_text
        assert all(
            worksheet.row_dimensions[row_index].height < 409
            for row_index in continuation_rows
        )


def test_reimports_only_annotation_fields_and_validates_final_csv(package_result):
    template = package_result.artifacts["pilot_annotator_1.csv"]
    workbook = load_workbook(io.BytesIO(build_annotation_workbook(template)))
    first_offer_sheet = workbook["Offre 01"]
    first_offer_sheet["B2"] = "Appellation technique modifiée"
    first_candidate_row = next(
        cell.row
        for cell in first_offer_sheet["C"]
        if isinstance(cell.value, str)
        and cell.value.startswith(PAIR_MARKER_PREFIX)
    )
    first_offer_sheet.cell(row=first_candidate_row + 1, column=2).value = (
        "RNCP technique modifié"
    )
    for worksheet in workbook.worksheets[1:]:
        for marker_cell in worksheet["C"]:
            if not isinstance(marker_cell.value, str) or not (
                marker_cell.value.startswith(ANNOTATION_MARKER_PREFIX)
            ):
                continue
            field, _, _ = marker_cell.value.removeprefix(
                ANNOTATION_MARKER_PREFIX
            ).partition(":")
            worksheet.cell(
                row=marker_cell.row,
                column=2,
                value={
                    "score": 2,
                    "critere_principal": VALID_MAIN_CRITERIA[0],
                    "justification": (
                        "Justification sur les compétences et la sécurité."
                    ),
                    "incertain": "NON",
                }[field],
            )
    output = io.BytesIO()
    workbook.save(output)

    completed_csv = import_workbook_annotations(template, output.getvalue())
    validation = validate_completed_annotation(template, completed_csv)
    template_rows = read_csv(template)
    completed_rows = read_csv(completed_csv)

    assert completed_csv.startswith(b"\xef\xbb\xbf")
    assert validation.pair_count == 60
    assert [row["pair_id"] for row in completed_rows] == [
        row["pair_id"] for row in template_rows
    ]
    for template_row, completed_row in zip(template_rows, completed_rows):
        for column in PROTECTED_COLUMNS:
            assert completed_row[column] == template_row[column]
        assert completed_row["score"] == "2"
        assert completed_row["critere_principal"] == VALID_MAIN_CRITERIA[0]
        assert completed_row["justification"] == (
            "Justification sur les compétences et la sécurité."
        )
        assert completed_row["incertain"] == "NON"


@pytest.mark.parametrize("invalid_pair_marker", ("missing", "unknown", "duplicate"))
def test_reimport_rejects_invalid_pair_markers(
    package_result,
    invalid_pair_marker,
):
    template = package_result.artifacts["pilot_annotator_1.csv"]
    workbook = load_workbook(io.BytesIO(build_annotation_workbook(template)))
    pair_cells = [
        cell
        for worksheet in workbook.worksheets[1:]
        for cell in worksheet["C"]
        if isinstance(cell.value, str)
        and cell.value.startswith(PAIR_MARKER_PREFIX)
    ]
    if invalid_pair_marker == "missing":
        pair_cells[0].value = None
    elif invalid_pair_marker == "unknown":
        pair_cells[0].value = PAIR_MARKER_PREFIX + "pair_inconnu"
    else:
        pair_cells[1].value = pair_cells[0].value
    output = io.BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="pair_id|réponses"):
        import_workbook_annotations(template, output.getvalue())


def test_export_adds_six_xlsx_companions_without_changing_csv_artifacts(
    package_result,
    tmp_path,
):
    export_annotation_packages(package_result, tmp_path)

    expected_workbooks = {
        f"{name.removesuffix('.csv')}.xlsx"
        for name in package_result.artifacts
        if name.endswith(".csv")
    }
    assert expected_workbooks == {
        "pilot_annotator_1.xlsx",
        "pilot_annotator_2.xlsx",
        "development_annotator_1.xlsx",
        "development_annotator_2.xlsx",
        "validation_annotator_1.xlsx",
        "validation_annotator_2.xlsx",
    }
    assert expected_workbooks <= {path.name for path in tmp_path.iterdir()}
    for name, content in package_result.artifacts.items():
        assert (tmp_path / name).read_bytes() == content


def test_manifest_contains_input_hashes_seeds_counts_and_audit_rules(package_result):
    manifest = json.loads(
        package_result.artifacts["annotation_manifest.json"].decode("utf-8")
    )

    assert len(manifest["source_lot5"]["sha256"]) == 4
    assert set(manifest["graines"]) == {"pilote", "annotator_1", "annotator_2"}
    assert manifest["paquets"]["pilot_annotator_1.csv"] == {
        "offres": 5,
        "couples": 60,
    }
    assert manifest["audit_doublons"]["quasi_doublon"][
        "jaccard_mots_minimal"
    ] == 0.95


def test_loads_and_cross_checks_the_four_lot5_artifacts():
    inputs = make_inputs()
    template_output = io.StringIO(newline="")
    writer = csv.DictWriter(
        template_output,
        fieldnames=("source", "source_offer_id", "code_rncp"),
        lineterminator="\n",
    )
    writer.writeheader()
    writer.writerows(
        {"source": source, "source_offer_id": offer_id, "code_rncp": code}
        for source, offer_id, code in sorted(inputs.annotation_pairs)
    )
    artifact_contents = {
        "sample_manifest.json": json.dumps(inputs.manifest).encode("utf-8"),
        "evaluation_offers.json": json.dumps(
            {"offres": inputs.offers}
        ).encode("utf-8"),
        "candidate_pools.jsonl": (
            "\n".join(json.dumps(pool) for pool in inputs.candidate_pools) + "\n"
        ).encode("utf-8"),
        "annotation_template.csv": template_output.getvalue().encode("utf-8"),
    }

    loaded = load_lot5_annotation_inputs(artifact_contents)

    assert len(loaded.offers) == 60
    assert len(loaded.annotation_pairs) == 720
    assert set(loaded.artifact_sha256) == set(artifact_contents)


def test_supports_generic_sources_and_non_standard_artifact_sizes():
    inputs = make_small_multisource_inputs()
    result = build_annotation_packages(inputs)
    reversed_result = build_annotation_packages(
        Lot5AnnotationInputs(
            manifest=inputs.manifest,
            offers=tuple(reversed(inputs.offers)),
            candidate_pools=tuple(reversed(inputs.candidate_pools)),
            annotation_pairs=inputs.annotation_pairs,
            artifact_sha256=inputs.artifact_sha256,
        )
    )
    development_rows = read_csv(
        result.artifacts["development_annotator_1.csv"]
    )
    validation_rows = read_csv(result.artifacts["validation_annotator_1.csv"])
    without_context_group = build_offer_group_id("FREE_WORK", "FW-001")
    generic_context_group = build_offer_group_id("FREE_WORK", "FW-002")
    without_context_rows = [
        row for row in development_rows if row["offer_group_id"] == without_context_group
    ]
    generic_context_rows = [
        row for row in development_rows if row["offer_group_id"] == generic_context_group
    ]

    assert result.pool_size == 3
    assert result.packet_counts["pilot_annotator_1.csv"] == {
        "offres": 5,
        "couples": 15,
    }
    assert result.packet_counts["development_annotator_1.csv"] == {
        "offres": 5,
        "couples": 15,
    }
    assert result.packet_counts["validation_annotator_1.csv"] == {
        "offres": 2,
        "couples": 6,
    }
    assert len({row["pair_id"] for row in development_rows}) == 15
    assert without_context_group != generic_context_group
    assert result.artifacts == reversed_result.artifacts
    assert all(row["offre_appellation"] == "" for row in without_context_rows)
    assert all(row["offre_code_rome"] == "" for row in without_context_rows)
    assert all(row["offre_libelle_rome"] == "" for row in without_context_rows)
    assert all(
        row["contexte_source_complementaire"] == ""
        for row in without_context_rows
    )
    assert all(
        row["contexte_source_complementaire"] == "Contexte Free-Work fourni."
        for row in generic_context_rows
    )
    assert "source" not in development_rows[0]
    completed_rows = complete_rows(
        result.artifacts["development_annotator_1.csv"]
    )
    validate_completed_annotation(
        result.artifacts["development_annotator_1.csv"],
        write_csv(completed_rows),
    )
    completed_rows[0]["offre_code_rome"] = "M1800"
    with pytest.raises(ValueError, match="Colonne de contexte modifiée"):
        validate_completed_annotation(
            result.artifacts["development_annotator_1.csv"],
            write_csv(completed_rows),
        )


def test_validates_a_completed_annotation_file(package_result):
    template = package_result.artifacts["pilot_annotator_1.csv"]
    completed = write_csv(complete_rows(template))

    result = validate_completed_annotation(template, completed)

    assert result.row_count == 60
    assert result.pair_count == 60


@pytest.mark.parametrize(
    ("mutation", "message"),
    (
        (lambda rows: rows[0].__setitem__("score", "4"), "Score invalide"),
        (
            lambda rows: rows[0].__setitem__("critere_principal", "INCONNU"),
            "Critère principal invalide",
        ),
        (lambda rows: rows[0].__setitem__("justification", ""), "Justification vide"),
        (
            lambda rows: rows[0].__setitem__("offre_intitule", "Texte modifié"),
            "Colonne de contexte modifiée",
        ),
    ),
)
def test_rejects_invalid_scores_criteria_justifications_and_context(
    package_result,
    mutation,
    message,
):
    template = package_result.artifacts["pilot_annotator_1.csv"]
    rows = complete_rows(template)
    mutation(rows)

    with pytest.raises(ValueError, match=message):
        validate_completed_annotation(template, write_csv(rows))


@pytest.mark.parametrize("change", ("missing", "duplicate"))
def test_rejects_removed_or_duplicated_rows(package_result, change):
    template = package_result.artifacts["pilot_annotator_1.csv"]
    rows = complete_rows(template)
    rows = rows[:-1] if change == "missing" else [*rows, dict(rows[0])]

    with pytest.raises(ValueError, match="pair_id"):
        validate_completed_annotation(template, write_csv(rows))
